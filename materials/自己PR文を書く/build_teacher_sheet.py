# -*- coding: utf-8 -*-
"""教員用評価シート.xlsx を生成するスクリプト"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

FONT_NAME = "游ゴシック"

NAVY = PatternFill("solid", fgColor="1F3864")
BLUE = PatternFill("solid", fgColor="D6E4F0")
GRAY = PatternFill("solid", fgColor="F2F2F2")
INPUT = PatternFill("solid", fgColor="FFFFF0")
PURPLE_F = PatternFill("solid", fgColor="E8DAEF")
GREEN_F = PatternFill("solid", fgColor="D5F5E3")
EXAMPLE_F = PatternFill("solid", fgColor="FFF9C4")

wrap = Alignment(wrap_text=True, vertical="top")
thin = Side(style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

UNIT_GOAL = ("自分の体験（具体）とそこから得た力や成長（抽象）とのつながりを整理し、"
             "高校入試の自己評価資料という目的に応じて、伝えたいことを明確にした自己PR文を"
             "書くことができる。")


def mg(ws, r1, c1, r2, c2, value, font=None, fill=None, align=None, border=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1)
    cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = align if align else wrap
    if border:
        for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
            for c in row:
                c.border = border
    return cell


wb = Workbook()

# ============================================================
# Sheet1: 回答入力
# ============================================================
ws1 = wb.active
ws1.title = "回答入力"
ws1.sheet_view.showGridLines = False

HEADERS = ["クラス", "出席番号", "氏名", "発問①\n（体験）", "発問②\n（学び）",
           "自己PR文\n（まとめ活動）", "AI評価\n(S〜D)", "AIコメント",
           "教師最終評価\n(S〜D)", "自己評価\n(S〜D)", "評価差考察", "振り返り"]
WIDTHS = [8, 8, 10, 26, 26, 32, 8, 30, 10, 8, 24, 24]

for i, w in enumerate(WIDTHS, start=1):
    ws1.column_dimensions[ws1.cell(row=1, column=i).column_letter].width = w

mg(ws1, 1, 1, 1, 12, "自己PR文を書く ―回答入力・評価シート―",
   Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF"), NAVY,
   Alignment(horizontal="center", vertical="center"))
ws1.row_dimensions[1].height = 26

mg(ws1, 2, 1, 2, 12, "単元目標：" + UNIT_GOAL,
   Font(name=FONT_NAME, size=10, bold=True), BLUE, wrap)
ws1.row_dimensions[2].height = 30

for col, h in enumerate(HEADERS, start=1):
    c = ws1.cell(row=3, column=col, value=h)
    c.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    c.fill = NAVY
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = box
ws1.row_dimensions[3].height = 32

DATA_START, DATA_END = 4, 43
dv_grade = DataValidation(type="list", formula1='"S,A,B,C,D"', allow_blank=True)
ws1.add_data_validation(dv_grade)

for r in range(DATA_START, DATA_END + 1):
    for col in range(1, 13):
        c = ws1.cell(row=r, column=col)
        c.border = box
        c.font = Font(name=FONT_NAME, size=10)
        c.alignment = wrap
        if col in (7, 9, 10):  # G:AI評価 I:教師評価 J:自己評価
            c.fill = PURPLE_F if col == 7 else (GREEN_F if col == 10 else GRAY)
        elif col in (4, 5, 6, 8, 11, 12):
            c.fill = INPUT
    ws1.row_dimensions[r].height = 30
    dv_grade.add(ws1.cell(row=r, column=7))
    dv_grade.add(ws1.cell(row=r, column=9))
    dv_grade.add(ws1.cell(row=r, column=10))

# 記入例（1行のみ、実データではないことが分かるよう色で明示）
example_row = [
    "3", "12", "山田 太郎",
    "文化祭のクラス合唱でパートリーダーを務めた。練習初日に声がまとまらず雰囲気が悪くなった。",
    "問題の本質を見極めて、小さく分解して向き合う力が身についたと感じている。",
    "私は文化祭のクラス合唱で、苦手な音楽のパートリーダーを務めました。練習初日、声がまとまら"
    "ずクラスの空気が重くなったとき、逃げずに一人ひとりの音を聞いて回り、苦手な部分だけを取り出"
    "して繰り返し練習する方法を提案しました。この経験から、私は「問題の本質を見極めて、小さく分"
    "解して向き合う力」が身についたと感じています。高校でも、難しい課題ほど一度立ち止まって整理"
    "してから取り組みたいです。",
    "S", "体験と学びの両方が具体的で、独自の言葉で表現されている。",
    "S", "S", "AI評価と自己評価が一致した。", "自分の言葉で書けたことが自信になった。",
]
for col, v in enumerate(example_row, start=1):
    c = ws1.cell(row=DATA_START, column=col, value=v)
    c.fill = EXAMPLE_F
    c.font = Font(name=FONT_NAME, size=9, italic=True, color="7D6608")

ws1.freeze_panes = "A4"

print("Sheet1 done")

# ============================================================
# Sheet2: AI設定
# ============================================================
ws2 = wb.create_sheet("AI設定")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 50

mg(ws2, 1, 1, 1, 2, "AI設定", Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF"), NAVY,
   Alignment(horizontal="center", vertical="center"))
ws2.row_dimensions[1].height = 26
ws2.row_dimensions[2].height = 8

settings = [
    (3, "AIモデル", "gemini-2.0-flash"),
    (4, "Temperature", 0.1),
    (6, "評価対象", "自己PR文（回答入力シートF列）"),
    (7, "知識・技能", "3段階（A/B/C）"),
    (8, "思考・判断・表現", "5段階（S/A/B/C/D）"),
    (10, "データ範囲", "回答入力シート 4行目〜43行目"),
    (11, "ClassroomフォルダID", "（ここにフォルダIDを入力）"),
]
for row, label, value in settings:
    c1 = ws2.cell(row=row, column=1, value=label)
    c1.font = Font(name=FONT_NAME, size=11, bold=True)
    c1.fill = GRAY
    c1.border = box
    c1.alignment = Alignment(vertical="center")
    c2 = ws2.cell(row=row, column=2, value=value)
    c2.font = Font(name=FONT_NAME, size=11)
    c2.fill = INPUT
    c2.border = box
    c2.alignment = Alignment(vertical="center", wrap_text=True)
    ws2.row_dimensions[row].height = 20

mg(ws2, 13, 1, 15, 2,
   "※ Gemini APIキーは、このシートには入力しません。GASの「プロジェクトの設定」→"
   "「スクリプト プロパティ」で GEMINI_API_KEY を設定してください。",
   Font(name=FONT_NAME, size=9.5, italic=True, color="C00000"), None, wrap)
for i in range(3):
    ws2.row_dimensions[13 + i].height = 18

print("Sheet2 done")

wb.save("/home/user/fohg/materials/自己PR文を書く/教員用評価シート.xlsx")
print("SAVED")
