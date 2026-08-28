# -*- coding: utf-8 -*-
"""生徒用ワークシート.xlsx を生成するスクリプト"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT_NAME = "游ゴシック"

# カラーパレット
NAVY = PatternFill("solid", fgColor="1F3864")
BLUE = PatternFill("solid", fgColor="D6E4F0")
LIGHT_BLUE = PatternFill("solid", fgColor="E8F0FE")
ORANGE_F = PatternFill("solid", fgColor="FFF2CC")
PURPLE_F = PatternFill("solid", fgColor="E8DAEF")
GREEN_F = PatternFill("solid", fgColor="D5F5E3")
GRAY = PatternFill("solid", fgColor="F2F2F2")
INPUT = PatternFill("solid", fgColor="FFFFF0")
AI_TIP = PatternFill("solid", fgColor="EBF5FB")
RED_FONT = Font(name=FONT_NAME, size=11, color="C00000", bold=True)

wrap = Alignment(wrap_text=True, vertical="top")
wrap_center = Alignment(wrap_text=True, vertical="center")
thin = Side(style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def mg(ws, r1, c1, r2, c2, value, font=None, fill=None, align=None, border=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1)
    cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = align if align else wrap
    if border:
        for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
            for c in row:
                c.border = border
    return cell


def section_header(ws, r, text, fill, span=(1, 4), font_color="FFFFFF", size=12):
    mg(ws, r, span[0], r, span[1], text,
       Font(name=FONT_NAME, size=size, bold=True, color=font_color), fill,
       Alignment(vertical="center", horizontal="left", indent=1))
    ws.row_dimensions[r].height = 24
    return r + 1


def ai_tip_row(ws, r, text, span=(1, 3)):
    mg(ws, r, span[0], r, span[1], "🤖 " + text,
       Font(name=FONT_NAME, size=9, color="1A5276", italic=True), AI_TIP, wrap)
    lines = text.count("\n") + 1 + (len(text) // 40)
    ws.row_dimensions[r].height = max(30, 14 * lines)
    return r + 1


def label_row(ws, r, text, span=(1, 3), height=20, bold=True, fill=GRAY, size=11):
    mg(ws, r, span[0], r, span[1], text,
       Font(name=FONT_NAME, size=size, bold=bold), fill, wrap)
    ws.row_dimensions[r].height = height
    return r + 1


def input_row(ws, r, span=(1, 3), height=90, counter_target=None, counter_max=None):
    mg(ws, r, span[0], r, span[1], None, Font(name=FONT_NAME, size=11), INPUT, wrap, box)
    ws.row_dimensions[r].height = height
    if counter_target is not None and counter_max is not None:
        cell_ref = f"{get_column_letter(span[0])}{r}"
        d_cell = ws.cell(row=r, column=span[1] + 1)
        d_cell.value = f'=IF({cell_ref}="","0/{counter_max}字",LEN({cell_ref})&"/{counter_max}字")'
        d_cell.font = Font(name=FONT_NAME, size=9, color="808080")
        d_cell.alignment = Alignment(vertical="top", horizontal="left")
    return r + 1


UNIT_GOAL = ("自分の体験（具体）とそこから得た力や成長（抽象）とのつながりを整理し、"
             "高校入試の自己評価資料という目的に応じて、伝えたいことを明確にした自己PR文を"
             "書くことができる。")

wb = Workbook()

# ============================================================
# Sheet1: 単元ガイド
# ============================================================
ws1 = wb.active
ws1.title = "単元ガイド"
ws1.sheet_view.showGridLines = False
for col, w in zip("ABCDEF", [10, 10, 12, 14, 14, 14]):
    ws1.column_dimensions[col].width = w

mg(ws1, 1, 1, 1, 6, "自分の歩みを言葉にする ―自己PR文を書こう―",
   Font(name=FONT_NAME, size=16, bold=True, color="1F3864"), None,
   Alignment(horizontal="center", vertical="center"))
ws1.row_dimensions[1].height = 32

# 基本情報欄
ws1["A2"] = 0
ws1["A2"].number_format = '0"組"'
ws1["A2"].fill = LIGHT_BLUE
ws1["A2"].font = Font(name=FONT_NAME, size=12)
ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws1["A2"].border = box

ws1["B2"] = 0
ws1["B2"].number_format = '0"番"'
ws1["B2"].fill = LIGHT_BLUE
ws1["B2"].font = Font(name=FONT_NAME, size=12)
ws1["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws1["B2"].border = box

ws1["C2"] = "名前"
ws1["C2"].font = Font(name=FONT_NAME, size=11, bold=True)
ws1["C2"].alignment = Alignment(horizontal="center", vertical="center")

mg(ws1, 2, 4, 2, 6, None, Font(name=FONT_NAME, size=12), LIGHT_BLUE,
   Alignment(horizontal="left", vertical="center", indent=1), box)
ws1.row_dimensions[2].height = 24

dv_class = DataValidation(type="whole", operator="between", formula1=1, formula2=20)
dv_number = DataValidation(type="whole", operator="between", formula1=1, formula2=50)
ws1.add_data_validation(dv_class)
ws1.add_data_validation(dv_number)
dv_class.add(ws1["A2"])
dv_number.add(ws1["B2"])

mg(ws1, 3, 1, 3, 6, "※ 組・番は半角数字で入力してね（例：2、15）", RED_FONT, None,
   Alignment(horizontal="left", vertical="center"))
ws1.row_dimensions[3].height = 18

ws1.row_dimensions[4].height = 8

r = section_header(ws1, 5, "■ 単元目標", NAVY, span=(1, 6))
mg(ws1, r, 1, r + 1, 6, UNIT_GOAL, Font(name=FONT_NAME, size=12, bold=True), BLUE, wrap)
ws1.row_dimensions[r].height = 24
ws1.row_dimensions[r + 1].height = 24
r += 2
ws1.row_dimensions[r].height = 8
r += 1

r = section_header(ws1, r, "■ この単元について", NAVY, span=(1, 6))
about_text = (
    "令和9年度埼玉県公立高等学校入学者選抜から、すべての受検生に「面接」が実施されることになりまし"
    "た。出願のときには「自己評価資料」を提出します。そこには、これまでの自分の体験を振り返り、"
    "力を注いだことや努力したこと、高校入学後や将来取り組んでみたいこと、自己PRなどを、自分の言葉"
    "で書きます。\n"
    "県の資料には、「実績そのものではなく、そこに至るまでの過程（プロセス）や意欲、身に付いた力な"
    "どを多面的に評価する」「文章・文字の巧拙は評価の対象外」「内容に正解はない」と書かれています。"
    "立派な実績を並べる必要はありません。小さな出来事でも、そこから自分が何を感じ、何を学んだかを"
    "自分の言葉で伝えることが大切です。この単元では、そのための自己PR文の書き方を学びます。"
)
mg(ws1, r, 1, r + 5, 6, about_text, Font(name=FONT_NAME, size=10.5), None, wrap)
for i in range(6):
    ws1.row_dimensions[r + i].height = 24
r += 6
ws1.row_dimensions[r].height = 8
r += 1

r = section_header(ws1, r, "■ 授業の流れ", NAVY, span=(1, 6))
schedule = [
    ("第1時", "① 自己PR文とは何かを知る　② 発問①：印象に残る体験を具体的に書く"
              "　③ 発問②：その体験から得た学びを自分の言葉でまとめる　④ AIと壁打ちして加筆する"),
    ("第2時", "① 自己PR文の条件を確認する　② 構想メモを作る　③ AIと壁打ちして構成を確かめる"
              "　④ 自己PR文（200〜300字）を書く　⑤ AIに添削してもらい推敲する"
              "　⑥ AI評価と自己評価を比べて振り返る"),
]
for title, detail in schedule:
    mg(ws1, r, 1, r, 2, title, Font(name=FONT_NAME, size=11, bold=True), GRAY,
       Alignment(horizontal="center", vertical="center"))
    mg(ws1, r, 3, r, 6, detail, Font(name=FONT_NAME, size=10.5), None, wrap)
    ws1.row_dimensions[r].height = 48
    r += 1
ws1.row_dimensions[r].height = 8
r += 1

r = section_header(ws1, r, "■ AIを使うときの約束", NAVY, span=(1, 6))
rules = (
    "・自分の体験や考えは、まず自分の力で書いてみよう。AIに最初から書いてもらわないこと。\n"
    "・AIには「書いてもらう」のではなく、「意見やヒントをもらう」ために使おう。\n"
    "・AIの言葉をそのまま貼り付けるのではなく、自分の言葉に直してから使おう。\n"
    "・自己評価資料には、生成AIの回答をそのまま転記・模倣してはいけません（自分で考え、記載すること）。"
)
mg(ws1, r, 1, r + 3, 6, rules, Font(name=FONT_NAME, size=10.5), None, wrap)
for i in range(4):
    ws1.row_dimensions[r + i].height = 22

print("Sheet1 done, last row:", r)

# ============================================================
# Sheet2: 評価基準
# ============================================================
ws2 = wb.create_sheet("評価基準")
ws2.sheet_view.showGridLines = False
for col, w in zip("ABCDE", [8, 30, 40, 40, 40]):
    ws2.column_dimensions[col].width = w

mg(ws2, 1, 1, 1, 5, "評価基準（ルーブリック）", Font(name=FONT_NAME, size=15, bold=True, color="1F3864"),
   None, Alignment(horizontal="center", vertical="center"))
ws2.row_dimensions[1].height = 28
mg(ws2, 2, 1, 2, 5, UNIT_GOAL, Font(name=FONT_NAME, size=10.5, bold=True), BLUE, wrap)
ws2.row_dimensions[2].height = 24
ws2.row_dimensions[3].height = 8

r = 4
r = section_header(ws2, r, "■ 知識・技能（3段階）", NAVY, span=(1, 5))
mg(ws2, r, 1, r, 5,
   "評価規準：体験（具体）と、そこから得た学び（抽象）とを、適切な語句を選んで結びつけて表現している"
   "か。",
   Font(name=FONT_NAME, size=10.5, italic=True), None, wrap)
ws2.row_dimensions[r].height = 22
r += 1

chishiki = [
    ("A", "体験と学びを的確な語句で結びつけ、読み手に伝わるように表現している。", GREEN_F),
    ("B", "体験と学びを書いているが、語句の選択や結びつけ方がやや曖昧である。", ORANGE_F),
    ("C", "体験または学びのどちらか一方のみの記述にとどまる、あるいは両者のつながりが読み取れない。",
     GRAY),
]
for label, text, fill in chishiki:
    mg(ws2, r, 1, r, 1, label, Font(name=FONT_NAME, size=13, bold=True), fill,
       Alignment(horizontal="center", vertical="center"), box)
    mg(ws2, r, 2, r, 5, text, Font(name=FONT_NAME, size=10.5), None, wrap, box)
    ws2.row_dimensions[r].height = 34
    r += 1
ws2.row_dimensions[r].height = 8
r += 1

r = section_header(ws2, r, "■ 思考・判断・表現（5段階）", NAVY, span=(1, 5))
mg(ws2, r, 1, r, 5,
   "評価規準：目的（高校入試の自己評価資料）に応じて、社会生活の中から題材を決め、伝えたいことを"
   "明確にして書いているか。",
   Font(name=FONT_NAME, size=10.5, italic=True), None, wrap)
ws2.row_dimensions[r].height = 22
r += 1

shihan = [
    ("S", "具体的な体験と、そこから得た学びの両方が、独自性・説得力をもって明確に書かれている。",
     PatternFill("solid", fgColor="F9E79F")),
    ("A", "具体的な体験と、そこから得た学びの両方が明確に書かれている。", GREEN_F),
    ("B", "具体的な体験、学びのどちらか一方のみが書かれている。", ORANGE_F),
    ("C", "体験も学びも抽象的・一般的な表現にとどまっている。", GRAY),
    ("D", "題材が定まっておらず、自己PR文として内容が成立していない。",
     PatternFill("solid", fgColor="F2D7D5")),
]
for label, text, fill in shihan:
    mg(ws2, r, 1, r, 1, label, Font(name=FONT_NAME, size=13, bold=True), fill,
       Alignment(horizontal="center", vertical="center"), box)
    mg(ws2, r, 2, r, 5, text, Font(name=FONT_NAME, size=10.5), None, wrap, box)
    ws2.row_dimensions[r].height = 34
    r += 1
ws2.row_dimensions[r].height = 8
r += 1

mg(ws2, r, 1, r + 2, 5,
   "★ B と A の分かれ目\n"
   "必須条件①「具体的な体験」と必須条件②「そこから得た学び」の【両方】に触れていなければ、A以上には"
   "なりません。片方だけならB、両方とも曖昧・欠けていればC、題材が定まっていなければDです。",
   Font(name=FONT_NAME, size=11, bold=True, color="C00000"),
   PatternFill("solid", fgColor="FDEDEC"), wrap, box)
for i in range(3):
    ws2.row_dimensions[r + i].height = 22

print("Sheet2 done")

# ============================================================
# Sheet3: 発問・回答
# ============================================================
ws3 = wb.create_sheet("発問・回答")
ws3.sheet_view.showGridLines = False
for col, w in zip("ABCD", [16, 16, 16, 18]):
    ws3.column_dimensions[col].width = w

mg(ws3, 1, 1, 1, 3, "発問・回答", Font(name=FONT_NAME, size=15, bold=True, color="1F3864"),
   None, Alignment(horizontal="center", vertical="center"))
ws3.row_dimensions[1].height = 28
mg(ws3, 2, 1, 2, 3, UNIT_GOAL, Font(name=FONT_NAME, size=10, bold=True), BLUE, wrap)
ws3.row_dimensions[2].height = 30
ws3.row_dimensions[3].height = 8

r = 4
r = section_header(ws3, r, "■ 第1時：体験を掘り起こそう", BLUE, span=(1, 3), font_color="1F3864")
ws3.row_dimensions[r].height = 6
r += 1

# --- 発問① ---
Q1_TEXT = ("発問① 体験の具体化\n"
           "あなたがこれまでの中学校生活（部活動・委員会・学校行事・学習・地域活動・家庭でのこと"
           "など、何でもよい）の中で、最も「力を注いだ」「努力した」と言える出来事を一つ選ぼう。その"
           "とき、どんな壁や難しさがあったか、いつ・どこで・誰と・何をしたのかが分かるように、具体的"
           "な場面を書こう。")
r = label_row(ws3, r, Q1_TEXT, height=110)
Q1_CELL_ROW = r
r = input_row(ws3, r, height=100, counter_target=True, counter_max=200)
r = ai_tip_row(ws3, r,
               "まず自分の体験を思い出して書いてみよう。書けたら「AIプロンプト」シートの壁打ち①に"
               "この内容を貼り付け、AIから「具体的で伝わりやすいか」意見をもらおう（書き直しは自分で"
               "行うこと）。")
ws3.row_dimensions[r].height = 6
r += 1

# --- 発問② ---
Q2_TEXT = ("発問② 具体から抽象への言語化\n"
           "発問①で書いた体験を振り返り、その出来事を通して自分がどう変わったか、どんな力が身につい"
           "たと感じるかを、自分だけの言葉で一文にまとめよう。「成長した」「頑張った」のような一般的"
           "な言葉ではなく、自分にしか言えない表現を探して書こう。")
r = label_row(ws3, r, Q2_TEXT, height=90)
Q2_CELL_ROW = r
r = input_row(ws3, r, height=80, counter_target=True, counter_max=150)
r = ai_tip_row(ws3, r,
               "発問①②の内容をまとめてAIプロンプトシートの壁打ち①に貼り付け、フィードバックをもら"
               "おう。AIの意見はあくまでヒント。書き直すかどうかは自分で判断しよう。")
ws3.row_dimensions[r].height = 10
r += 1

r = section_header(ws3, r, "■ 第2時：自己PR文を書こう", ORANGE_F, span=(1, 3), font_color="7D6608")
ws3.row_dimensions[r].height = 6
r += 1

mg(ws3, r, 1, r, 3,
   "場面設定：あなたが志望する高校に提出する「自己評価資料」の自己PR欄に書くつもりで書こう。",
   RED_FONT, None, wrap)
ws3.row_dimensions[r].height = 30
r += 1

mg(ws3, r, 1, r, 3,
   "必須条件　① 具体的な体験（エピソード）が書かれていること　"
   "② その体験から得た学び・成長した力が、体験と結びつけて書かれていること　"
   "字数：200〜300字",
   Font(name=FONT_NAME, size=10.5, italic=True), None, wrap)
ws3.row_dimensions[r].height = 34
r += 1

r = label_row(ws3, r, "構想メモ（体験→学び→高校生活への意欲を簡単に整理しよう）", height=20)
r = input_row(ws3, r, height=70)
r = ai_tip_row(ws3, r,
               "構想メモができたら、AIプロンプトシートの壁打ち②に貼り付け、必須条件①②が満たされて"
               "いそうかアドバイスをもらおう。AIに代わりに書いてもらうのはNG。")
ws3.row_dimensions[r].height = 6
r += 1

r = label_row(ws3, r, "自己PR文（200〜300字）", height=20, fill=ORANGE_F)
ESSAY_CELL_ROW = r
r = input_row(ws3, r, height=130, counter_target=True, counter_max=300)
r = ai_tip_row(ws3, r,
               "書き終えたら、AIプロンプトシートの添削プロンプトに貼り付け、分かりやすさや具体性に"
               "ついてフィードバックをもらおう。AIに書き直してもらうのではなく、自分の言葉で推敲する"
               "こと。")
ws3.row_dimensions[r].height = 10
r += 1

r = section_header(ws3, r, "■ AI評価と自己評価", PURPLE_F, span=(1, 3), font_color="4A235A")
ws3.row_dimensions[r].height = 6
r += 1

mg(ws3, r, 1, r, 2, "AI評価（S〜D）", Font(name=FONT_NAME, size=11, bold=True), GRAY, wrap)
AI_EVAL_ROW = r
c = ws3.cell(row=r, column=3)
c.fill = PURPLE_F
c.border = box
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[r].height = 22
r += 1

mg(ws3, r, 1, r, 2, "自己評価（S〜D）", Font(name=FONT_NAME, size=11, bold=True), GRAY, wrap)
SELF_EVAL_ROW = r
c = ws3.cell(row=r, column=3)
c.fill = GREEN_F
c.border = box
c.alignment = Alignment(horizontal="center", vertical="center")
dv_grade = DataValidation(type="list", formula1='"S,A,B,C,D"', allow_blank=True)
ws3.add_data_validation(dv_grade)
dv_grade.add(c)
ws3.row_dimensions[r].height = 22
r += 1

r = ai_tip_row(ws3, r,
               "AI評価と自己評価を比べてみよう。同じだった？　違った場合、なぜその違いが生まれたのか"
               "を考えることが、自分の書き方の癖に気づく学びになるよ。")
ws3.row_dimensions[r].height = 6
r += 1

r = label_row(ws3, r, "評価差考察（AI評価と自己評価の違い・気づいたこと）", height=20)
DIFF_ROW = r
r = input_row(ws3, r, height=60)
ws3.row_dimensions[r].height = 10
r += 1

r = section_header(ws3, r, "■ 振り返り", GREEN_F, span=(1, 3), font_color="196F3D")
ws3.row_dimensions[r].height = 6
r += 1
r = label_row(ws3, r, "この単元を通して考えたこと・学んだことを書こう", height=20)
REFLECT_ROW = r
r = input_row(ws3, r, height=70)

print("Sheet3 done. Cell map:")
print("Q1_CELL =", f"A{Q1_CELL_ROW}")
print("Q2_CELL =", f"A{Q2_CELL_ROW}")
print("ESSAY_CELL =", f"A{ESSAY_CELL_ROW}")
print("AI_EVAL_CELL =", f"C{AI_EVAL_ROW}")
print("SELF_EVAL_CELL =", f"C{SELF_EVAL_ROW}")
print("DIFF_CELL =", f"A{DIFF_ROW}")
print("REFLECT_CELL =", f"A{REFLECT_ROW}")

# ============================================================
# Sheet4: AIプロンプト
# ============================================================
ws4 = wb.create_sheet("AIプロンプト")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 95

mg(ws4, 1, 1, 1, 1, "AIプロンプト集（コピペして使おう）",
   Font(name=FONT_NAME, size=15, bold=True, color="1F3864"), None,
   Alignment(horizontal="center", vertical="center"))
ws4.row_dimensions[1].height = 28
mg(ws4, 2, 1, 2, 1,
   "使い方：それぞれの見出しの下にあるプロンプトを丸ごとコピーし、AIチャットに貼り付けて、"
   "指示にある「★貼り付け箇所★」に自分が書いた文章を入れて使おう。",
   Font(name=FONT_NAME, size=10.5, italic=True), None, wrap)
ws4.row_dimensions[2].height = 34
ws4.row_dimensions[3].height = 8

r = 4
r = section_header(ws4, r, "① 壁打ち①（発問チェック）― 第1時で使う", BLUE, span=(1, 1),
                    font_color="1F3864")

prompt1 = (
    "あなたは中学3年生の作文相談にのる先生です。中学3年生にわかりやすい言葉で答えてください。\n"
    "以下は、高校入試の自己PR文を書くために、私が考えた「体験」と「そこから得た学び」です。\n\n"
    "【体験】\n★ここに発問①の答えを貼り付け★\n\n"
    "【そこから得た学び】\n★ここに発問②の答えを貼り付け★\n\n"
    "次の２点についてアドバイスをください。書き直した文章は書かず、ヒントだけを300字以内で"
    "教えてください。\n"
    "1. 体験の場面が具体的でわかりやすいか\n"
    "2. 学びの内容が自分の言葉で表現できているか"
)
mg(ws4, r, 1, r, 1, prompt1, Font(name=FONT_NAME, size=10.5), INPUT, wrap, box)
ws4.row_dimensions[r].height = 220
r += 1
ws4.row_dimensions[r].height = 12
r += 1

r = section_header(ws4, r, "② 壁打ち②（構想チェック）― 第2時の前半で使う", ORANGE_F, span=(1, 1),
                    font_color="7D6608")

prompt2 = (
    "あなたは中学3年生の作文相談にのる先生です。中学3年生にわかりやすい言葉で答えてください。\n"
    "私は、高校入試で提出する「自己評価資料」の自己PR欄（200〜300字）を書こうとしています。\n"
    "場面：志望する高校に提出する自己評価資料の自己PR欄\n"
    "必須条件：① 具体的な体験（エピソード）が書かれていること　② その体験から得た学び・成長した力"
    "が、体験と結びつけて書かれていること\n\n"
    "【私の構想メモ】\n★ここに構想メモを貼り付け★\n\n"
    "この構想メモが、①②の必須条件を満たせそうかチェックし、構成についてのアドバイスだけを300字"
    "以内で教えてください。代わりに文章を書くことはしないでください。"
)
mg(ws4, r, 1, r, 1, prompt2, Font(name=FONT_NAME, size=10.5), INPUT, wrap, box)
ws4.row_dimensions[r].height = 220
r += 1
ws4.row_dimensions[r].height = 12
r += 1

r = section_header(ws4, r, "③ 添削（まとめ活動アドバイス）― 第2時の後半で使う", PURPLE_F, span=(1, 1),
                    font_color="4A235A")

prompt3 = (
    "あなたは中学3年生の作文相談にのる先生です。中学3年生にわかりやすい言葉で答えてください。\n"
    "以下は、私が書いた高校入試の自己PR文（自己評価資料の自己PR欄、200〜300字）です。\n\n"
    "【私の自己PR文】\n★ここに自己PR文を貼り付け★\n\n"
    "次のチェックリストを確認し、書き直した文章は書かず、ヒントだけを400字以内で教えてください。\n"
    "□ 具体的な体験（エピソード）が書かれているか\n"
    "□ その体験から得た学び・成長した力が、体験と結びつけて書かれているか\n"
    "□ 200〜300字に収まっているか\n"
    "□ 読み手（高校の先生）に伝わる言葉になっているか"
)
mg(ws4, r, 1, r, 1, prompt3, Font(name=FONT_NAME, size=10.5), INPUT, wrap, box)
ws4.row_dimensions[r].height = 230

print("Sheet4 done")

wb.save("/home/user/fohg/materials/自己PR文を書く/生徒用ワークシート.xlsx")
print("SAVED")
